from selenium import webdriver
from bs4 import BeautifulSoup
from selenium.webdriver.common.by import By
import requests
from lxml import etree
import pandas as pd
import re
import time
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.desired_capabilities import DesiredCapabilities
import PySimpleGUI as sg
import requests
from openpyxl import Workbook
from openpyxl import load_workbook
from selenium.webdriver.chrome.options import Options
import keyboard
import pickle

# cookies = {
#     'timeOffset_o': '-2466.89990234375',
#     'smidV2': '20240213131526eda6a072fd197c5595f8ccf08e74636300697a88373e3e480',
#     'token': 'NzEzNjUwMDR8ZDRmMjE4Yzc4ZDMzMDZhMjVjYjlhNTI0MGNkMTc3MGR8fHx8MjU5MjAwMHwxfHx85pmL5rGf55So5oi3fDB8bW9iaWxlfDF8MHx8',
#     'testcookie': 'yes',
#     'JJSESS': '%7B%22clicktype%22%3A%22%22%7D',
#     'Hm_lvt_bc3b748c21fe5cf393d26c12b2c38d99': '1707872801',
#     'Hm_lpvt_bc3b748c21fe5cf393d26c12b2c38d99': '1707873731',
#     'JJEVER': '%7B%22shumeideviceId%22%3A%22WHJMrwNw1k/HzkbeyzLLlTzBmlATQhqOtodsedjQbgbQWB918uk823M6hyK7PMpzySN+FsNH+sa3SDnT2ECVL2MvabL1UxbYOdCW1tldyDzmQI99+chXEik1CT1Q++Cne27A54gDgTBk3vaAoz2wRe0lBLaZv6y0KU10Rt18csPbSczD/UEAZLhaM0sfKvSdcbh4vSnbcMzY2bCBM9cOgFa/Z7ELUEgHVIK252guvG/Hl6AdSbHxuXwUDic6AW2yIixSMW2OQhYo%3D1487582755342%22%2C%22fenzhan%22%3A%22dm%22%2C%22fenpin%22%3A%22gbl.html%22%2C%22nicknameAndsign%22%3A%222%257E%2529%2524yan54%22%2C%22foreverreader%22%3A%2271365004%22%2C%22desid%22%3A%22YRJ96LU3B+kVY9psf/LLM/y6tJNAzAzv%22%2C%22sms_total%22%3A9%2C%22lastCheckLoginTimePc%22%3A1707887890%7D',
# }

# headers = {
#     'Accept': 'text/javascript, application/javascript, application/ecmascript, application/x-ecmascript, */*; q=0.01',
#     'Accept-Language': 'zh-CN,zh;q=0.9',
#     'Connection': 'keep-alive',
#     # 'Cookie': 'timeOffset_o=-2466.89990234375; smidV2=20240213131526eda6a072fd197c5595f8ccf08e74636300697a88373e3e480; token=NzEzNjUwMDR8ZDRmMjE4Yzc4ZDMzMDZhMjVjYjlhNTI0MGNkMTc3MGR8fHx8MjU5MjAwMHwxfHx85pmL5rGf55So5oi3fDB8bW9iaWxlfDF8MHx8; testcookie=yes; JJSESS=%7B%22clicktype%22%3A%22%22%7D; Hm_lvt_bc3b748c21fe5cf393d26c12b2c38d99=1707872801; Hm_lpvt_bc3b748c21fe5cf393d26c12b2c38d99=1707873731; JJEVER=%7B%22shumeideviceId%22%3A%22WHJMrwNw1k/HzkbeyzLLlTzBmlATQhqOtodsedjQbgbQWB918uk823M6hyK7PMpzySN+FsNH+sa3SDnT2ECVL2MvabL1UxbYOdCW1tldyDzmQI99+chXEik1CT1Q++Cne27A54gDgTBk3vaAoz2wRe0lBLaZv6y0KU10Rt18csPbSczD/UEAZLhaM0sfKvSdcbh4vSnbcMzY2bCBM9cOgFa/Z7ELUEgHVIK252guvG/Hl6AdSbHxuXwUDic6AW2yIixSMW2OQhYo%3D1487582755342%22%2C%22fenzhan%22%3A%22dm%22%2C%22fenpin%22%3A%22gbl.html%22%2C%22nicknameAndsign%22%3A%222%257E%2529%2524yan54%22%2C%22foreverreader%22%3A%2271365004%22%2C%22desid%22%3A%22YRJ96LU3B+kVY9psf/LLM/y6tJNAzAzv%22%2C%22sms_total%22%3A9%2C%22lastCheckLoginTimePc%22%3A1707887890%7D',
#     'Referer': 'https://www.jjwxc.net/onebook.php?novelid=6442002',
#     'Sec-Fetch-Dest': 'empty',
#     'Sec-Fetch-Mode': 'cors',
#     'Sec-Fetch-Site': 'same-origin',
#     'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/121.0.0.0 Safari/537.36',
#     'X-Requested-With': 'XMLHttpRequest',
#     'sec-ch-ua': '"Chromium";v="121", "Not A(Brand";v="99"',
#     'sec-ch-ua-mobile': '?0',
#     'sec-ch-ua-platform': '"Windows"',
# }

# params = {
#     'callback': 'novelreviewCallback',
#     'action': 'getByNovelid',
#     'novelid': '6442002',
# }

# response = requests.get('https://www.jjwxc.net/novelreview.php', params=params, cookies=cookies, headers=headers)


def request(hre,match,all2,m,matchs,running):
    
    
        
    
        # 检测键盘按键
    
    
        if keyboard.is_pressed('q'):
            running = False
            print('已停止')
        else:
            # opts = Options()
            # opts.headless = True  # 设置无头模式，相当于执行了opt.add_argument('--headless')和opt.add_argument('--disable-gpu')(--disable-gpu禁用gpu加速仅windows系统下执行)。
            # browser = webdriver.Chrome(options=opts)  # 如果没有将chromedriver加入环境变量，第一个参数需传入其绝对路径
            
            #第一步输入这个：去除开头警告
           
    
    
            #控制台打印
            print("开始执行你的测试用例！")
    
    
            #第二个输入这个：隐藏式启动谷歌浏览器执行UI测试用例
            chrome_options = Options()
            chrome_options.add_argument('--headless')
            browser= webdriver.Chrome(options=chrome_options)

            # browser = webdriver.Chrome()
            browser.implicitly_wait(10)
            browser.get(hre)
            
            # browser.execute_script('scrollTo(0,1000000)')
            time.sleep(3)
            html=browser.page_source
            # f=open('./2.html',mode="w",encoding="utf-8")
            # f.write(html)
            # # 读取HTML文件
            # with open('./2.html', mode='r', encoding='utf-8') as f:
            #     html = f.read()
            # print(html)
            html =etree.HTML(html)
            try:
                
                jianjie= html.xpath('//*[@id="novelintro"]/text()')
                jianjie = ' '.join(jianjie)
                jianjie =jianjie.replace('\u3000\u3000','').replace('\u3000','')
            # print(jianjie)
            except:
                jianjie =[]
                print('1')
        
            try:
                labs= html.xpath("//div[@class='smallreadbody']/span/text()")
                lable =[]
                for lab in labs:
                    lab = lab.strip('\n').strip(' ').rstrip('\n') .strip()   
                    lable.append(lab)
            except:
                lable=[]
                print('2')
            try:
                titles= html.xpath('//span/h1/span/text()|//*[@id="oneboolt"]/div[1]/h1/a/span/text()')
                title=[]
                for ti in titles:
                    ti=ti.strip('[]').strip(' ')
                    title.append(ti)
            except:
                title=[]
                print('3')
            # print(title)
            try:
                
                author= html.xpath('//div/h2/a/span/text()|//*[@id="oneboolt"]/div[1]/span/a/text()')
            except:
                author=[]
                print(4)
            try:
                
                others= html.xpath("//td[@class='sptd']/div[@align='center']/text()")[4]
                jif = re.findall(r'文章积分：([\d,]+)', others)
                wz = ''.join(jif)
            except:
                print(5)
                ot = html.xpath("//tbody/tr/td[1]/div[@style='padding-top: 50px;']/text()")
                wz=[]
                for w in ot:
                    jif = re.findall(r'文章积分：([\d,]+)', w)
                    wz.append(jif)
                wz = ''.join(jif)
                # print(wz)
                
            # print(others)
            try:
                others= html.xpath("//td[@class='sptd']/div[@align='center']/text()")[4]
                click= html.xpath("//td[@class='sptd']/div[@align='center']/span[1]/text()")

                
                sp= html.xpath("//td[@class='sptd']/div[@align='center']/span[2]/text()")
                shoucang= html.xpath("//td[@class='sptd']/div[@align='center']/span[3]/text()")
                liuip= html.xpath("//td[@class='sptd']/div[@align='center']/span[4]/text()")
                # print(click)
                # print(sp)
                # print(shoucang)
            except:
                print(6)
                
                othe = html.xpath("//tbody/tr/td[1]/div[@style='padding-top: 50px;']/text()")
                print(othe)
                click=[]
                liuip=[]
                
                sp=[]
                shoucang=[]
                
                for w in othe:
                    oth = re.findall(r'(\d+)', w)
                    print(oth)
                    click.append(oth[0])
                    
                    sp.append(oth[1])
                    shoucang.append(oth[2])
                click
                print(click)
                print(sp) 
                print(shoucang)
                    
                    
                #     w.strip('\n').strip(' ').rstrip('\n').strip('\u3000').strip()
                #     shuping = re.findall(r'总书评数：([\d,]+)', w)
                #     sp.append(shuping)
                #     shoucangs = re.findall(r'当前被收藏数：([\d,]+)', w)
                #     shoucang.append(shoucangs)
                #     dianji =re.findall(r'总点评数：(\d+)', w)
                #     click.append(dianji)
                #     sp = ''.join(shuping)
                #     # print(sp)
                #     shoucang = ''.join(shoucangs)
                
                # print(click)
                # print(shoucang)
                    
                    
            # try:        
            #     click = ''.join(dianji)
            #     print(click)
            


            #     # 分割字符串，获取结果数组的第二个元素
            #     click_number = click.split('：')[1].strip()
            #     print(click_number)

            #     # print(click_number)  # 输出：10720
            #     # # stripped_data = data.strip()
            #     # # print(stripped_data)
            #     # # datas.append(stripped_data)
            #     # # print(datas)

            #     click = re.findall(r'(\d+)', click_number)[0]


                
            #     print(click)
            #     # sp = ''.join(shuping)
            #     # print(sp)
            #     # shoucang = ''.join(shoucangs)
                
                
            #     # print(shoucang)
                
                
            # except:
            #     print(click)
            #     click =click[0]
            #     print(click)
                
                
            #     print('555')  
                
            
                
                
            try:    
                shijia=html.xpath("/html/body/table[1]/tbody/tr/td/div/ul/li[2]/text()")
                shijiao=[]
                for s in shijia:
                    s=s.strip('\n').strip(' ').rstrip('\n').strip()
                    if s:
                        shijiao.append(s)
                
                xilies =html.xpath("/html/body/table[1]/tbody/tr/td/div/ul/li[3]/span[2]/text()")
                xilie=[]
                for xi in xilies:
                    xi = xi.strip('\n').strip(' ').rstrip('\n').strip('\xa0').strip()
                    xilie.append(xi)
                    
                jindu =html.xpath("/html/body/table[1]/tbody/tr/td/div/ul/li[4]/span[2]/text() | /html/body/table[1]/tbody/tr/td/div/ul/li[4]/span/font/text()")
                
                
                
                zishu =html.xpath("/html/body/table[1]/tbody/tr/td/div/ul/li[5]/span[2]/text() ")
                chus =html.xpath("/html/body/table[1]/tbody/tr/td/div/ul/li[6]/text()|/html/body/table[1]/tbody/tr/td/div/ul/li[6]/img/@alt")
                chuban=[]
                for chu in chus:
                    chu=chu.strip('\n').strip(' ').rstrip('\n').strip()
                    if chu:
                        chuban.append(chu)
                
                        
                rongyu =html.xpath("/html/body/table[1]/tbody/tr/td/div/ul/li[8]/div[1]/p[1]/text()|/html/body/table[1]/tbody/tr/td/div/ul/li[8]/div[2]/p/text()")
            except:
                shijiao=[]
                xilie=[]
                jindu=[]
                zishu=[]
                chus=[]
                chuban=[]
                print(7)
            try:
                rank= html.xpath('/html/body/table[1]/tbody/tr/td[3]/div[4]/div[1]/div[1]/span[1]/text()')
            except:
                print(8)
            book=[]
            book2={}
            match =match
            # bs4(match,book2,book)
            # for al in all2:
            book2['类型']=all2[m-1]['类型']
            book2['时间']=all2[m-1]['时间']
            
            book2['文章标题']=title
            book2['作者']=author
            
            book2['文章简介']=jianjie
            book2['文章积分']=wz
            book2['非v点击数']=click
            book2['霸王票排名']=rank
            book2['总书评数']=sp
            book2['当前被收藏数']=shoucang
            book2['营养液数']=liuip
            book2['作品视角']=shijiao
            book2['所属系列']=xilie
            book2['进度']=jindu
            book2['字数']=zishu
            book2['出版情况']=chuban
            book2['荣誉']=rongyu
            bs4(match,book2,book)
            
            book2['书编号']=matchs
            
            # if html.xpath("html/body/table/tbody/tr/td/div[@class='righttd' and @id='novelreview_div']/div/span/text()"):
            
            # pingfen =html.xpath("html/body/table/tbody/tr/td/div[@class='righttd' and @id='novelreview_div']/div/span/text()")
            # print(pingfen)
                # book2['总评分']=pingfen[0]
                # book2['评分人数']=pingfen[1]
                # book2['评分比例']=pingfen[2]
                # book.append(book2)
            book.append(book2)
            
            book3 = [{k: v[0].strip() if isinstance(v, list) and len(v) == 1 else v for k, v in item.items()} for item in book]
            # 将data1数据写入Excel的Sheet1
            # writer = pd.ExcelWriter('data.xlsx', engine='openpyxl')
            # df1 = pd.DataFrame(book3)
            
            # df1.to_excel(writer, sheet_name='Sheet4')
            workbook = Workbook()
            # if m==1:
                
            #     # 将data1数据写入Excel的Sheet1
            #     df = pd.DataFrame(book3)
            #     writer = pd.ExcelWriter	('晋江.xlsx', engine='openpyxl')
            #     writer.book = workbook
            #     df.to_excel(writer, sheet_name='作者', index=False)
            #     writer.save()
            # else:
            try:
                book = load_workbook('晋江.xlsx')
                writer = pd.ExcelWriter('晋江.xlsx', engine='openpyxl')
                writer.book = book

                df_new = pd.DataFrame(book3)

                # 选择要追加数据的工作表，并追加数据
                sheet_name = '作品'
                start_row = writer.sheets[sheet_name].max_row
                df_new.to_excel(writer, sheet_name=sheet_name, index=False, startrow=start_row, header=False)

                writer.save()
            except:
                print('未保存')
                
            
                

        

        
        
        



            print(book3)
            return book3
def selenium():
    
    browser = webdriver.Chrome()
    browser.get('https://www.jjwxc.net/comment.php?novelid=5484954&wonderful=1')
    browser.implicitly_wait(10)
    comment = browser.find_elements(By.XPATH, '/html/body/table/tbody/tr/td/div/div/div/div[2]/span[1]')
    comments =[]
    for com in comment:
      
        
        temp ={}
        
        temp["title"] = com.text
        comments.append(temp)
        print(comments)
        
def bs4(match,book2,book):
    
    url="https://www.jjwxc.net/novelreview.php?callback=novelreviewCallback&action=getByNovelid&{}".format(match)
    cookies={"Cookie":"timeOffset_o=-2466.89990234375; smidV2=20240213131526eda6a072fd197c5595f8ccf08e74636300697a88373e3e480; token=NzEzNjUwMDR8ZDRmMjE4Yzc4ZDMzMDZhMjVjYjlhNTI0MGNkMTc3MGR8fHx8MjU5MjAwMHwxfHx85pmL5rGf55So5oi3fDB8bW9iaWxlfDF8MHx8; testcookie=yes; JJSESS=%7B%22clicktype%22%3A%22%22%7D; Hm_lvt_bc3b748c21fe5cf393d26c12b2c38d99=1707872801; Hm_lpvt_bc3b748c21fe5cf393d26c12b2c38d99=1707873731; JJEVER=%7B%22shumeideviceId%22%3A%22WHJMrwNw1k/HzkbeyzLLlTzBmlATQhqOtodsedjQbgbQWB918uk823M6hyK7PMpzySN+FsNH+sa3SDnT2ECVL2MvabL1UxbYOdCW1tldyDzmQI99+chXEik1CT1Q++Cne27A54gDgTBk3vaAoz2wRe0lBLaZv6y0KU10Rt18csPbSczD/UEAZLhaM0sfKvSdcbh4vSnbcMzY2bCBM9cOgFa/Z7ELUEgHVIK252guvG/Hl6AdSbHxuXwUDic6AW2yIixSMW2OQhYo%3D1487582755342%22%2C%22fenzhan%22%3A%22dm%22%2C%22fenpin%22%3A%22gbl.html%22%2C%22nicknameAndsign%22%3A%222%257E%2529%2524yan54%22%2C%22foreverreader%22%3A%2271365004%22%2C%22desid%22%3A%22YRJ96LU3B+kVY9psf/LLM/y6tJNAzAzv%22%2C%22sms_total%22%3A9%2C%22lastCheckLoginTimePc%22%3A1707887890%7D"}
    headers= {'User-Agent':
'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/121.0.0.0 Safari/537.36'}
    response=requests.get(url,cookies=cookies,headers=headers)
    print(response)
    res = response.text
    print('正在bs4')
    if len(res)<=400:
        book2['已评人数']='暂无'
        book2['平均评分']='暂无'
        book2['已评比例']='暂无'
        
    if len(res)>400:
        start_pos = res.find("({") + 1
        end_pos = res.rfind("})") + 1
        data = res[start_pos:end_pos]
        data_dict = eval(data)
        # try:
            
        #     datas1 = data_dict["data"]["scoreData"]["novelid"]
        #     book2['书编号']=datas1
        # except:
            
        #     book2['书编号']=''
        #     print('有问题1')
        try:
            
            datas2 = data_dict["data"]["scoreData"]["num"]
            book2['已评人数']=datas2
        except:
            book2['已评人数']='暂无'
            print('有问题2')
        try:
            
            datas3 = data_dict["data"]["scoreData"]["avgscore"]
            book2['评均评分']=datas3
        except:
            book2['平均评分']='暂无'
            print('有问题3')
        try:
            
            datas4 = data_dict["data"]["scoreData"]["percent"]
            book2['已评比例']=datas4
        except:
            book2['已评比例']='暂无'
            print('有问题4')
            
            
            
        
        
    
        
        
    # print(datas)
    
            
def reader(counter,running,o):
    
    if keyboard.is_pressed('q'):
        running = False
        print("已退出1")
    
    

    else:
         #控制台打印
        print("开始执行你的测试用例！")


        #第二个输入这个：隐藏式启动谷歌浏览器执行UI测试用例
        chrome_options = Options()
        chrome_options.add_argument('--headless')
        browser= webdriver.Chrome(options=chrome_options)

        # browser = webdriver.Chrome()
        browser.get('https://www.jjwxc.net/onereader.php?readerid={}'.format(counter))
        
        browser.implicitly_wait(10)
        html=browser.page_source
        # f=open('./4.html',mode="w",encoding="utf-8")
        # f.write(html)
        # # 读取HTML文件
        # with open('./4.html', mode='r', encoding='utf-8') as f:
        #     html = f.read()
        
        html =etree.HTML(html) 
        # zhuangtai=html.xpath().click()
        # duzhe={}
        m=html.xpath('//table/tbody/tr/td/div/a')
        
        if m:
            
            duzhe_all =[]
            duzhe_all2=[]
            duzhe_all3=[]
            try:
                name=html.xpath('/html/body/table/tbody/tr/td[1]/table/tbody/tr[1]/td/span[1]/font/b/text()')
            except:
                print("01")
                name=[]
            try:
                pingjunfeng = html.xpath('//*[@id="novelreview"]/div/div/div[1]/span[2]/text()')
            except:
                print("02")
                pingjunfeng=[]
            try:
                yiping = html.xpath('//*[@id="novelreview"]/div/div/div[1]/span[3]/text()')
            except:
                print("03")
                yiping=[]
            try:
                beipingbi= html.xpath('//*[@id="novelreview"]/div/div/div[1]/span[4]/text()')
            except:
                print("04")
                beipingbi=[]
            try:
                
                youpingfenqianxian = html.xpath('//*[@id="novelreview"]/div/div/div[1]/span[5]/text()')
            except:
                print("05")
                youpingfenqianxian=[]
            try:
                # 评论
                pinglunzuoping =html.xpath('//*[@id="novelreview"]/div/table/tbody[1]/tr/td[1]/a/text()')
                # print(pinglunzuoping)
                
                pinglunzuoping_hrefs=html.xpath('//*[@id="novelreview"]/div/table/tbody[1]/tr/td[1]/a/@href')
                pinglunzuoping_href=[]
                for pinglunzuoping_hre in pinglunzuoping_hrefs:
                    
                    match = re.search( r"novelid=(\d+)", pinglunzuoping_hre)
                    match=match.group()
                    
                    pinglunzuoping_href.append(match)
                pinglurank=html.xpath('//*[@id="novelreview"]/div/table/tbody[1]/tr[1]/td[2]/a/span/text()')
            except:
                print("06")
                pinglunzuoping=[]
                pinglunzuoping_href=[]
                pinglurank=[]
            for p,f,n in zip(pinglunzuoping,pinglunzuoping_href,pinglurank):
                duzhe={}
                duzhe['名字']=name
                
                duzhe['评论的作品']=p
                
                duzhe['id']=f
                
                duzhe['评论等级']=n
                duzhe['读者id']=str(counter)
                
                duzhe_all2.append(duzhe)
                
            # print(pinglunzuoping_href)
            
                
                
            try:    
                
                # 霸王榜
                author =html.xpath('//*[@id="ticketList"]/dl/dt/a/text()')
                # print(author)
                author_id =html.xpath('//*[@id="ticketList"]/dl/dt/a/@href')
                author_ids=[]
                for id in author_id:
                    
                    match = re.search( r"(\d+)", id)
                    match=match.group()
                    
                    author_ids.append(match)
                # print(author_ids)
                rank =html.xpath('//*[@id="ticketList"]/dl/dd/span/text()')
                # print(rank)
            except:
                print("07")
                author=[]
                author_ids=[]
                rank=[]
            for i,m,n in zip(author,author_ids,rank):
                al={}
                al['名字']=name
                al["霸王榜的作者名字"]=i
                al['作者id']=m
                al['等级']=n
                al['读者id']=str(counter)
                duzhe_all3.append(al)
            try:
            
                book_hrefs =html.xpath('//table/tbody/tr/td/div/a/@href')
                book_names =html.xpath('//table/tbody/tr/td/div/a/text()')
                book_name=[]
                # book_href=[]
            except:
                print("08")
                book_hrefs=[]
                book_names=[]
                book_name=[]
                
            
            for book,hre in zip(book_names,book_hrefs):
                duzhe={}
                book=book.strip('\n').strip(' ').rstrip('\n').replace('\xa0', '')
                if book:
                    book_name.append(book)
                    duzhe['名字']=name
                    duzhe['收藏的书名']=book
                    
                
                match = re.search( r"(\d+)", hre)
                match=match.group()
                # book_href.append(match)
                duzhe['编码']=match
                duzhe['平均分']=pingjunfeng
                duzhe['已评作品']=yiping
                duzhe['被屏蔽评论']=beipingbi
                duzhe['有评分权限的作品']=youpingfenqianxian
                duzhe['读者id']=str(counter)
            
                # duzhe['评论的作品']=pinglunzuoping
                # duzhe['评论作品的id']=pinglunzuoping_href

                
                
                
                duzhe_all.append(duzhe)
                
            # for hre in book_hrefs:
                
                
            #     match = re.search( r"novelid=(\d+)", hre)
            #     match=match.group()
            #     book_href.append(match)
                
                # print(match)
                
            # print(book_name,book_href)
            # quan = browser.find_elements(By.XPATH, '//*[@id="load_show_novelsa"]/table/tbody/tr/td/div/a')
            
            # al = []
            # for i in quan:
            #     i=i.text
            #     i=i.strip('\n').strip(' ').rstrip('\n')
            #     al.append(i)
            # print(al)
            try:
                kaifang = browser.find_elements(By.XPATH, '//td[contains(@style,"cursor: pointer;")][@onclick]')
                for i in kaifang:
                
                    i.click() 
                    browser.switch_to.window(browser.window_handles[-1])
                    time.sleep(3)
                    kaifang = browser.find_elements(By.XPATH, '//table/tbody/tr/td/div/a')
                    href = browser.find_elements(By.XPATH, '//table/tbody/tr/td/div/a')
                    
                    for m,n in zip(kaifang,href):
                        m=m.text
                        if m not in book_name:
                            duzhe={}
                            href=n.get_attribute('href')
                            match = re.search( r"novelid=(\d+)", href)
                            match=match.group()
                            duzhe={}
                            duzhe['名字']=name
                            duzhe["收藏的书名"]=m
                            duzhe["编码"]=match
                            
                            duzhe['已评作品']=yiping
                            duzhe['被屏蔽评论']=beipingbi
                            duzhe['有评分权限的作品']=youpingfenqianxian
                            duzhe['读者id']=str(counter)
                            duzhe['平均分']=pingjunfeng
                            # duzhe['评论的作品']=pinglunzuoping
                            # duzhe['评论作品的id']=pinglunzuoping_href

                    
                            
                            duzhe_all.append(duzhe)
                        
                    
                
                    browser.close()
                    browser.switch_to.window(browser.window_handles[0])
            except:
                print("09")
                kaifang=[]
            # for i in kaifang:
                
            #     i.click() 
            #     browser.switch_to.window(browser.window_handles[-1])
            #     time.sleep(3)
            #     kaifang = browser.find_elements(By.XPATH, '//table/tbody/tr/td/div/a')
            #     href = browser.find_elements(By.XPATH, '//table/tbody/tr/td/div/a')
                
            #     for m,n in zip(kaifang,href):
            #         m=m.text
            #         if m not in book_name:
            #             duzhe={}
            #             href=n.get_attribute('href')
            #             match = re.search( r"novelid=(\d+)", href)
            #             match=match.group()
            #             duzhe={}
            #             duzhe['名字']=name
            #             duzhe["收藏的书名"]=m
            #             duzhe["编码"]=match
            #             duzhe['平均分']=pingjunfeng
            #             duzhe['已评作品']=yiping
            #             duzhe['被屏蔽评论']=beipingbi
            #             duzhe['有评分权限的作品']=youpingfenqianxian
            #             duzhe['读者id']=str(counter+1)
            #             # duzhe['评论的作品']=pinglunzuoping
            #             # duzhe['评论作品的id']=pinglunzuoping_href

                
                        
            #             duzhe_all.append(duzhe)
                    
                
            
            #     browser.close()
            #     browser.switch_to.window(browser.window_handles[0])
            
                # time.sleep(2)
            print(duzhe_all)
            print(duzhe_all2)
            print(duzhe_all3)
            duzhe_all = [{k: v[0].strip() if isinstance(v, list) and len(v) == 1 else v for k, v in item.items()} for item in duzhe_all]
            duzhe_all2 = [{k: v[0].strip() if isinstance(v, list) and len(v) == 1 else v for k, v in item.items()} for item in duzhe_all2]
            duzhe_all3 = [{k: v[0].strip() if isinstance(v, list) and len(v) == 1 else v for k, v in item.items()} for item in duzhe_all3]
            workbook = Workbook()
            # if m==1:
                
            #     # 将data1数据写入Excel的Sheet1
            #     df = pd.DataFrame(book3)
            #     writer = pd.ExcelWriter	('晋江.xlsx', engine='openpyxl')
            #     writer.book = workbook
            #     df.to_excel(writer, sheet_name='作者', index=False)
            #     writer.save()
            # else:
            book = load_workbook('晋江.xlsx')
            writer = pd.ExcelWriter('晋江.xlsx', engine='openpyxl')
            writer.book = book

            
            try:
                df_new = pd.DataFrame(duzhe_all)
                

                # 选择要追加数据的工作表，并追加数据
                sheet_name = '读者的基本信息'
                start_row = writer.sheets[sheet_name].max_row
                df_new.to_excel(writer, sheet_name=sheet_name, index=False, startrow=start_row, header=False)
                
            except:
                print('无读者基本信息')
            try:
                df_new = pd.DataFrame(duzhe_all2)
                
                # 选择要追加数据的工作表，并追加数据
                sheet_name = '读者点评的作品'
                start_row = writer.sheets[sheet_name].max_row
                df_new.to_excel(writer, sheet_name=sheet_name, index=False, startrow=start_row, header=False)

               
            except:
                print('无读者点评的作品')
            try:
                df_new = pd.DataFrame(duzhe_all3)

                # 选择要追加数据的工作表，并追加数据
                sheet_name = '读者霸王榜'
                start_row = writer.sheets[sheet_name].max_row
                df_new.to_excel(writer, sheet_name=sheet_name, index=False, startrow=start_row, header=False)
            except:
                print("无读者霸王榜")
                
            writer.save()
                        
                # df = pd.DataFrame(duzhe_all)
                # df.to_csv('data.csv', index=False)
                
                # book_hrefs =html.xpath('//table/tbody/tr/td/div/a/@href')
                # book_names =html.xpath('//table/tbody/tr/td/div/a/text()')
                # book_name=[]
                # book_href=[]
                # for book in book_names:
                #     book=book.strip('\n').strip(' ').rstrip('\n').replace('\xa0', '')
                #     if book:
                #         book_name.append(book)
                # for hre in book_hrefs:
                    
                    
                #     match = re.search( r"novelid=(\d+)", hre)
                #     match=match.group()
                #     book_href.append(match)
                    
                    # print(match)
                    
                # print(book_name,book_href)
                # return duzhe_all,duzhe_all2,duzhe_all3
                #   data_list =[job,salary,education,experience,location]
                #     get_data = pd.DataFrame(columns = ['职位名称',"薪资",'教育水平','经验','地点'])
                #     for col,data in zip(get_data.columns,data_list):
                #         get_data[col] =data
                    
                #     final_df_dict[num]=get_data
                #     print('第{}页爬取完成'.format(num+1))
                # concat_df=pd.concat(list(final_df_dict.values()),ignore_index=True)
                # file_name = '{}招聘信息.csv'.format(city)
                # path=r'C:\Users\yan\Desktop\计算机\爬虫\{}'.format(file_name)
                # concat_df.to_csv(file_name,encoding='utf-8',index=False)
                # print('{}保存成功'.format(file_name))
                # 创建一个Excel写入对象
                # writer = pd.ExcelWriter('data.xlsx', engine='openpyxl')

                # # 将data1数据写入Excel的Sheet1
                # df1 = pd.DataFrame(duzhe_all)
                # df1.to_excel(writer, sheet_name='Sheet1')

                # # 将data2数据写入Excel的Sheet2
                # df2 = pd.DataFrame(duzhe_all2)
                # df2.to_excel(writer, sheet_name='Sheet2')

                # # 将data3数据写入Excel的Sheet3
                # df3 = pd.DataFrame(duzhe_all3)
                # df3.to_excel(writer, sheet_name='Sheet3')

                # # 保存Excel文件
                # writer.save()
            # except:
            #     print('未保存/无数据')
            # running =False
            # o=o
            
def autuor(counter,running,o):


    

        # 检测键盘按键
        if keyboard.is_pressed('q'):
                running = False
    

        else:
            # opts = Options()
            # opts.headless = True  # 设置无头模式，相当于执行了opt.add_argument('--headless')和opt.add_argument('--disable-gpu')(--disable-gpu禁用gpu加速仅windows系统下执行)。
            # browser = webdriver.Chrome(options=opts)  # 如果没有将chromedriver加入环境变量，第一个参数需传入其绝对路径
         #控制台打印
            print("开始执行你的测试用例！")
    
    
            #第二个输入这个：隐藏式启动谷歌浏览器执行UI测试用例
            chrome_options = Options()
            chrome_options.add_argument('--headless')
            browser= webdriver.Chrome(options=chrome_options)
            # browser = webdriver.Chrome()
            browser.implicitly_wait(10)
            browser.get('https://www.jjwxc.net/oneauthor.php?authorid={}'.format(counter))
            
            html=browser.page_source
            # f=open('./3.html',mode="w",encoding="utf-8")
            # f.write(html)
            # # 读取HTML文件
            # with open('./3.html', mode='r', encoding='utf-8') as f:
            #     html = f.read()
            
            html =etree.HTML(html)
            hrefs = html.xpath('//table/tbody/tr/td/div[@style]/a/@href')
            if hrefs:
            
                all =[]
                # time.sleep(5)
                zuopin= html.xpath('//table/tbody/tr/td/div/a/text()')
                zuoping = []
                for zuo in zuopin:
                    
                    zuo = zuo.strip().strip('\u3000').split(' ')[-1].replace('\xa0', '')
                    # if zuo:
                    #     temp = {}
                    zuoping.append(zuo)
                    #     temp['作品']=zuo
                    
                        
                    # print(temp)
                    # all.append(temp)
                # print(all)
                zuoping = [text for text in zuoping if text.strip() != '']
                hrefs = html.xpath('//table/tbody/tr/td/div[@style]/a/@href')
                locks =html.xpath("//a[@class='tooltip' and @href]/text()|//img[@style='vertical-align:middle;']/../a/text()")
                lock=[]
                
                for loc in locks:
                    
                    loc = loc.strip().split(' ')[-1].replace('\xa0', '')
                    if loc:
                    #     temp = {}
                        lock.append(loc)
                print(lock)
                href =[]
                for hre in hrefs:
                    hre = 'https://www.jjwxc.net/' + hre
                    # print(hre)
                    # time.sleep(5)
                    href.append(hre)
                    # request(hre)
                    # if hre:
                    #     temp = {}
                    #     href.append(hre)
                    #     temp['href']=hre
                    #     all.append(temp)
                    
                # print(all)
                types = html.xpath('//table/tbody/tr/td[2]/text()')
                type =[]
                for ty in types:
                    typ=ty.strip('\n').strip(' ').rstrip('\n').replace('\xa0', '').replace('类型', '')
                    type.append(typ)
                
                type = [text for text in type if text.strip() != '']
                nums=html.xpath('//table/tbody/tr/td[4]/text()')
                number =[]
                for num in nums:
                    num=num.strip('\n').strip(' ').rstrip('\n').replace('\xa0', '').replace('字数', '')
                    number.append(num)
                number = [text for text in number if text.strip() != '']
                
                times =html.xpath('//tbody/tr/td[6]/text()')
                time=[]
                for tim in times:
                    tim =tim.replace('发表时间', '')
                    time.append(tim)
                time = [text for text in time if text.strip() != '']
                for z1,h1,t1,n1 in zip(zuoping,href,type,number):
                    
                    if str(n1) != '0':
                        all.append((z1,h1,t1,n1))
                    #     temp ={}
                    #     temp['作品'] = z1
                    #     temp['链接'] = h1
                    #     temp['类型'] = t1
                    #     temp['字数'] = n1
                        # for m1 in time:
                        #     temp['时间'] = m1
                        #     all.append(temp)
                            
                        # all.append(temp)
                        
                        # print(z1,h1,t1,n1)
                        # request(hre)
                # print(all)
                
                
                
                all1=[]
                
                for a1,m1 in zip(all,time):
                    temp ={}
                    temp['作品'] = a1[0]
                    temp['链接'] = a1[1]
                    temp['类型'] = a1[2]
                    temp['字数'] = a1[3]
                    temp['时间'] = m1
                    all1.append(temp)
                # print(all1)
                # print(lock)
                
                all2 = [item for item in all1 if item['作品'] not in lock]
                print(all2)
                m=0
                        
                for hre in all2:
                    hre =hre['链接']
                    
                    match = re.search( r"novelid=(\d+)", hre)
                    matchs =re.search( r"(\d+)", hre)
                    match=match.group()
                    matchs=matchs.group()
                    print(match)
                    m+=1
                    
                    request(hre,match,all2,m,matchs,running)
            
def save_counter_value(counter,o):
    with open('{}.pkl'.format(o), 'wb') as file:
        pickle.dump(counter, file)
    print('已保存编号')

def load_counter_value(o):
    try:
        with open('{}.pkl'.format(o), 'rb') as file:
            counter = pickle.load(file)
            return counter
    except FileNotFoundError:
        return 0

if __name__ == '__main__':
    print('提示，按q退不了就一直点/n')
    o=input('你想要读取(author,reader):')
    
    previous_counter = load_counter_value(o)
    print('上次保存的值是:', previous_counter)
    
    # 设置初始值
    counter = previous_counter
    running = True
    
    while running:
        counter += 1
        
        if o=="reader":
            
            
        
            print('当前的值:', counter)
            reader(counter,running,o)
        elif o=="author":
            
            
        
            print('当前的值:', counter)
            autuor(counter,running,o)
        else:
            print('啥？')
            break
        # 执行计数器的逻辑
    
        # counter += 1
        
        # print('当前的值:', counter)
        
    
        # 检测键盘按键
        if keyboard.is_pressed('q'):
            running = False
            print('已退出')
            
            
        
        
    
    # 在循环结束之后，保存当前的计数器值，以便下次使用
    save_counter_value(counter,o)
    